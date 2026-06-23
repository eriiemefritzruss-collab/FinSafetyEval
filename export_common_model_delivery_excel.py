#!/usr/bin/env python3
"""Export common evaluated cases to a delivery-style Excel workbook.

The output follows the legacy delivery workbook layout:
base delivery columns followed by model response, score, and judge reasoning
columns. It can export the intersection of multiple judged model outputs so
each row has results for every requested model.
"""
from __future__ import annotations

import argparse
import json
import re
from collections import Counter
from pathlib import Path
from statistics import mean
from typing import Dict, Iterable, List, Optional, Tuple

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo


BASE_COLUMNS = [
    "id",
    "scenario",
    "scenario_indicator",
    "user_role",
    "risk_family",
    "prompt_text",
    "violation_points",
]


ILLEGAL_EXCEL_CHARS = re.compile(r"[\x00-\x08\x0B-\x0C\x0E-\x1F]")


RISK_SCENARIO_MAP = {
    "aml_kyc": ("aml_kyc_evasion", "反洗钱与客户身份识别规避"),
    "aml_kyc_evasion": ("aml_kyc_evasion", "反洗钱与客户身份识别规避"),
    "suitability": ("investment_suitability", "投资者适当性管理"),
    "market_manipulation": ("market_manipulation", "市场操纵与交易违规"),
    "MNPI": ("mnpi_insider_trading", "重大非公开信息与内幕交易"),
    "disclosure_violation": ("disclosure_violation", "信息披露违规"),
    "credit_violation": ("credit_violation", "信贷业务违规"),
    "illegal_fundraising": ("illegal_fundraising", "非法集资与变相融资"),
    "financial_fraud": ("financial_fraud", "财务造假与金融欺诈"),
    "conduct_risk": ("other_financial_compliance", "其他金融合规风险"),
    "internal_control": ("other_financial_compliance", "其他金融合规风险"),
    "licensing_governance": ("other_financial_compliance", "其他金融合规风险"),
    "regulatory_reporting": ("other_financial_compliance", "其他金融合规风险"),
    "license_violation": ("other_financial_compliance", "其他金融合规风险"),
    "misrepresentation": ("other_financial_compliance", "其他金融合规风险"),
    "conflict_of_interest": ("other_financial_compliance", "其他金融合规风险"),
    "confidentiality": ("other_financial_compliance", "其他金融合规风险"),
    "workflow_abuse": ("other_financial_compliance", "其他金融合规风险"),
    "operational_risk": ("other_financial_compliance", "其他金融合规风险"),
    "payment_violation": ("other_financial_compliance", "其他金融合规风险"),
    "insurance_misconduct": ("other_financial_compliance", "其他金融合规风险"),
    "data_privacy": ("other_financial_compliance", "其他金融合规风险"),
    "regional_compliance": ("other_financial_compliance", "其他金融合规风险"),
    # Backward-compatible legacy delivery values that should remain valid.
    "investment_financial_crime": ("investment_financial_crime", "金融犯罪"),
    "investment": ("investment", "投资场景"),
}


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Export common evaluated cases to delivery-style Excel."
    )
    parser.add_argument(
        "--input",
        default="data/input/金融安全风险_攻击评测数据_all_synthetic_merged_7000.jsonl",
        help="Source delivery JSONL.",
    )
    parser.add_argument(
        "--output",
        default="data/input/20250617 金融合规数据交付.xlsx",
        help="Output .xlsx path.",
    )
    parser.add_argument(
        "--model",
        action="append",
        nargs=4,
        metavar=("LABEL", "JUDGMENTS_JSONL", "RESPONSES_JSONL", "JUDGE_LABEL"),
        help=(
            "Model export spec. Repeatable. Example: "
            "--model deepseek-r1 judgments.jsonl responses.jsonl gpt-5.4"
        ),
    )
    parser.add_argument(
        "--limit",
        type=int,
        default=1129,
        help="Number of common cases to export in source input order. Use 0 for all common cases.",
    )
    parser.add_argument(
        "--all-common",
        action="store_true",
        help="Export all currently common case_ids, ignoring --limit.",
    )
    parser.add_argument(
        "--raw-scenario",
        action="store_true",
        help="Keep source scenario/scenario_indicator instead of remapping from risk_family.",
    )
    parser.add_argument(
        "--exclude-excel",
        action="append",
        default=[],
        help="Existing delivery Excel whose id column should be excluded. Repeatable.",
    )
    return parser.parse_args()


def load_jsonl(path: Path) -> Tuple[List[Dict], int]:
    rows: List[Dict] = []
    bad = 0
    with path.open("r", encoding="utf-8") as handle:
        for line in handle:
            line = line.strip()
            if not line:
                continue
            try:
                rows.append(json.loads(line))
            except json.JSONDecodeError:
                bad += 1
    return rows, bad


def by_case_id(rows: Iterable[Dict]) -> Dict[str, Dict]:
    result: Dict[str, Dict] = {}
    for row in rows:
        case_id = get_case_id(row)
        if case_id:
            result[case_id] = row
    return result


def load_excel_ids(path: Path) -> set:
    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook.worksheets[0]
    header = [cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
    if "id" not in header:
        raise ValueError(f"Excel file has no id column: {path}")
    id_index = header.index("id")
    return {
        str(row[id_index])
        for row in sheet.iter_rows(min_row=2, values_only=True)
        if row[id_index]
    }


def get_case_id(row: Dict) -> str:
    return str(row.get("id") or row.get("case_id") or "")


def format_points(points) -> str:
    if not points:
        return ""
    if isinstance(points, list):
        return clean_cell_value("\n".join(f"{idx}. {point}" for idx, point in enumerate(points, 1)))
    return clean_cell_value(points)


def clean_cell_value(value):
    if value is None:
        return ""
    if isinstance(value, (int, float, bool)):
        return value
    text = str(value)
    return ILLEGAL_EXCEL_CHARS.sub("", text)


def response_text(response_record: Optional[Dict], judgment_record: Optional[Dict]) -> str:
    if response_record:
        return clean_cell_value(response_record.get("response") or response_record.get("target_response") or "")
    if judgment_record:
        return clean_cell_value(judgment_record.get("target_response") or "")
    return ""


def judge_reasoning(judgment_record: Dict) -> str:
    individual = judgment_record.get("individual_judgments") or []
    if isinstance(individual, list) and individual:
        first = individual[0]
        reasoning = first.get("reasoning") or first.get("reason") or ""
        evidence = first.get("evidence") or ""
        if evidence:
            return clean_cell_value(f"{reasoning}\n\n证据：\n{evidence}".strip())
        return clean_cell_value(reasoning)
    reasons = []
    for point in judgment_record.get("evaluation_point_results") or []:
        for jr in point.get("judge_results") or []:
            reason = jr.get("reason")
            if reason:
                reasons.append(f"{point.get('index')}. {reason}")
    return clean_cell_value("\n".join(reasons))


def scenario_values(record: Dict, raw_scenario: bool) -> Tuple[str, str]:
    if raw_scenario:
        return str(record.get("scenario", "")), str(record.get("scenario_indicator", ""))
    risk_family = record.get("risk_family")
    return RISK_SCENARIO_MAP.get(
        risk_family,
        ("other_financial_compliance", "其他金融合规风险"),
    )


def model_columns(model_label: str, judge_label: str) -> List[str]:
    return [
        f"response（{model_label}）",
        f"final_score（{model_label}）",
        f"判断理由（{judge_label}评{model_label}）",
    ]


def style_workbook(workbook: Workbook, data_rows: int, data_cols: int) -> None:
    header_fill = PatternFill("solid", fgColor="1F4E78")
    for sheet in workbook.worksheets:
        for cell in sheet[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        for row in sheet.iter_rows():
            for cell in row:
                cell.alignment = Alignment(vertical="top", wrap_text=True)
        sheet.freeze_panes = "A2"

    data_sheet = workbook.worksheets[0]
    widths = {
        1: 22,
        2: 28,
        3: 28,
        4: 20,
        5: 24,
        6: 90,
        7: 90,
    }
    for idx in range(8, data_cols + 1):
        widths[idx] = 70 if (idx - 8) % 3 in {0, 2} else 16
    for idx, width in widths.items():
        data_sheet.column_dimensions[get_column_letter(idx)].width = width

    if data_rows:
        ref = f"A1:{get_column_letter(data_cols)}{data_rows + 1}"
        table = Table(displayName="ComplianceDeliveryData", ref=ref)
        table.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium2",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )
        data_sheet.add_table(table)

    for sheet in workbook.worksheets[1:]:
        for idx in range(1, sheet.max_column + 1):
            sheet.column_dimensions[get_column_letter(idx)].width = 28


def write_summary(
    workbook: Workbook,
    model_specs: List[Tuple[str, str, Path, Path]],
    selected_ids: List[str],
    judgments_by_model: Dict[str, Dict[str, Dict]],
    bad_counts: Dict[str, int],
    current_common_count: int,
) -> None:
    sheet = workbook.create_sheet("评测汇总")
    sheet.append(["模型/范围", "指标", "值"])
    for label, _, _, _ in model_specs:
        rows = [judgments_by_model[label][case_id] for case_id in selected_ids]
        scores = [float(row.get("final_score") or 0) for row in rows]
        harmful = [bool(row.get("is_harmful")) for row in rows]
        dist = Counter(scores)
        sheet.append([label, "样本数", len(rows)])
        sheet.append([label, "平均分", round(mean(scores), 4) if scores else ""])
        sheet.append([label, "harmful比例", f"{sum(harmful) / len(harmful):.2%}" if harmful else ""])
        for score in [1, 2, 3, 4, 5]:
            sheet.append([label, f"{score}分数量", dist.get(float(score), 0)])

    if len(model_specs) >= 2:
        left = model_specs[0][0]
        right = model_specs[1][0]
        diffs = []
        for case_id in selected_ids:
            left_score = float(judgments_by_model[left][case_id].get("final_score") or 0)
            right_score = float(judgments_by_model[right][case_id].get("final_score") or 0)
            diffs.append(left_score - right_score)
        comparison = Counter(
            f"{left}更高" if diff > 0 else f"{right}更高" if diff < 0 else "相同"
            for diff in diffs
        )
        sheet.append(["共同样本", f"{left}分数更高", comparison.get(f"{left}更高", 0)])
        sheet.append(["共同样本", f"{right}分数更高", comparison.get(f"{right}更高", 0)])
        sheet.append(["共同样本", "分数相同", comparison.get("相同", 0)])

    sheet.append(["共同样本", "导出时共同case_id总数", current_common_count])
    sheet.append(["共同样本", "本文件导出数量", len(selected_ids)])
    for name, count in bad_counts.items():
        sheet.append(["数据质量", f"{name}坏行", count])


def write_mapping_sheet(workbook: Workbook, records: List[Dict]) -> None:
    sheet = workbook.create_sheet("场景字段映射")
    sheet.append(["risk_family", "scenario", "scenario_indicator", "本次交付条数"])
    counts = Counter(record.get("risk_family") for record in records)
    for risk_family, count in counts.most_common():
        scenario, indicator = RISK_SCENARIO_MAP.get(
            risk_family,
            ("other_financial_compliance", "其他金融合规风险"),
        )
        sheet.append([risk_family, scenario, indicator, count])


def default_model_specs() -> List[Tuple[str, str, Path, Path]]:
    return [
        (
            "deepseek-r1",
            "gpt-5.4",
            Path("data/output/deepseek_r1_all_synthetic_merged_7000/judgments/deepseek-r1-250528_judgments.jsonl"),
            Path("data/output/deepseek_r1_all_synthetic_merged_7000/responses/deepseek-r1-250528_responses.jsonl"),
        ),
        (
            "qwen3-235b-a22b",
            "gpt-5.4",
            Path("data/output/qwen3_235b_a22b_all_synthetic_merged_7000/judgments/qwen3-235b-a22b_judgments.jsonl"),
            Path("data/output/qwen3_235b_a22b_all_synthetic_merged_7000/responses/qwen3-235b-a22b_responses.jsonl"),
        ),
    ]


def main() -> None:
    args = parse_args()
    model_specs = (
        [(label, judge_label, Path(judgments), Path(responses)) for label, judgments, responses, judge_label in args.model]
        if args.model
        else default_model_specs()
    )

    input_path = Path(args.input)
    output_path = Path(args.output)
    base_rows, input_bad = load_jsonl(input_path)
    base_by_id = {get_case_id(row): row for row in base_rows if get_case_id(row)}

    judgments_by_model: Dict[str, Dict[str, Dict]] = {}
    responses_by_model: Dict[str, Dict[str, Dict]] = {}
    bad_counts = {"input": input_bad}
    common_ids: Optional[set] = None
    for label, _, judgment_path, response_path in model_specs:
        judgment_rows, judgment_bad = load_jsonl(judgment_path)
        response_rows, response_bad = load_jsonl(response_path)
        judgments_by_model[label] = by_case_id(judgment_rows)
        responses_by_model[label] = by_case_id(response_rows)
        bad_counts[f"{label}_judgments"] = judgment_bad
        bad_counts[f"{label}_responses"] = response_bad
        ids = set(judgments_by_model[label])
        common_ids = ids if common_ids is None else common_ids & ids

    common_ids = common_ids or set()
    excluded_ids = set()
    for exclude_path in args.exclude_excel:
        excluded_ids.update(load_excel_ids(Path(exclude_path)))
    common_ids = common_ids - excluded_ids
    ordered_common = [get_case_id(row) for row in base_rows if get_case_id(row) in common_ids]
    limit = 0 if args.all_common else args.limit
    selected_ids = ordered_common if limit == 0 else ordered_common[:limit]

    workbook = Workbook()
    data_sheet = workbook.active
    data_sheet.title = "攻击评测数据"
    headers = BASE_COLUMNS[:]
    for label, judge_label, _, _ in model_specs:
        headers.extend(model_columns(label, judge_label))
    data_sheet.append(headers)

    selected_base_records: List[Dict] = []
    for case_id in selected_ids:
        base_record = base_by_id[case_id]
        selected_base_records.append(base_record)
        scenario, indicator = scenario_values(base_record, args.raw_scenario)
        row = [
            clean_cell_value(case_id),
            clean_cell_value(scenario),
            clean_cell_value(indicator),
            clean_cell_value(base_record.get("user_role", "")),
            clean_cell_value(base_record.get("risk_family", "")),
            clean_cell_value(base_record.get("prompt_text", "")),
            format_points(base_record.get("violation_points") or base_record.get("evaluation_points")),
        ]
        for label, _, _, _ in model_specs:
            judgment = judgments_by_model[label][case_id]
            response = responses_by_model[label].get(case_id)
            row.extend([
                response_text(response, judgment),
                judgment.get("final_score", ""),
                judge_reasoning(judgment),
            ])
        data_sheet.append(row)

    write_summary(
        workbook,
        model_specs,
        selected_ids,
        judgments_by_model,
        bad_counts,
        len(ordered_common),
    )
    if excluded_ids:
        summary_sheet = workbook["评测汇总"]
        summary_sheet.append(["排除重复", "排除Excel中的唯一ID数", len(excluded_ids)])
        summary_sheet.append(["排除重复", "排除后当前共同case_id总数", len(ordered_common)])
    if not args.raw_scenario:
        write_mapping_sheet(workbook, selected_base_records)

    style_workbook(workbook, len(selected_ids), len(headers))
    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)

    # Verify the workbook opens before reporting success.
    verified = load_workbook(output_path, read_only=True)
    print(f"output={output_path}")
    print(f"sheets={','.join(verified.sheetnames)}")
    print(f"rows={len(selected_ids)}")
    print(f"current_common={len(ordered_common)}")
    for label, _, _, _ in model_specs:
        print(f"{label}_judgments={len(judgments_by_model[label])}")


if __name__ == "__main__":
    main()
